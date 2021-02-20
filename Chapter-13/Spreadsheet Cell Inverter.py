"""
Spreadsheet Cell Inverter.
Makes rows, columns and columns, rows.
"""

import openpyxl

# Opens and excel spreadsheet to convert
wb1 = openpyxl.load_workbook('before.xlsx')
ws1 = wb1['Sheet']

# Opens a destination spreadsheet
wb2 = openpyxl.load_workbook('after.xlsx')
ws2 = wb2.active

# Set ax row and max column variables
mr = ws1.max_row
mc = ws1.max_column

# Saves rows as columns and columns as rows, by switching j and i in between
# reading the original file and writing to the new file
for i in range(1, mc + 1):
    for j in range(1, mr + 1):
        copycell = ws1.cell(row=j, column=i).value
        ws2.cell(row=i, column=j).value = copycell

# Save the spreadsheet
wb2.save('after.xlsx')
