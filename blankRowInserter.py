"""
Blank Row Inserter.
Inserts blank row(s) into an Excel spreadsheet.
"""

import openpyxl

# Ask where to put row and how many
rowNum = int(input('Insert row number for the insertion of blank lines: '))
blankNum = int(input('How many rows? '))

# Opens a workbook, change the name.
wb1 = openpyxl.load_workbook('produceSalestext.xlsx')
ws1 = wb1['Sheet']

# Open a new workbook that will contain the data with inserted rows
wb2 = openpyxl.load_workbook('rowinserterProduceSales.xlsx')
ws2 = wb2.active

# Max row and column variables
mr = ws1.max_row
mc = ws1.max_column

# Loops through rows until the one where a row will be inserted, and copies
# to new file.
for i in range(1, rowNum):
    for j in range(1, mc + 1):
        copycell = ws1.cell(row=i, column=j).value
        ws2.cell(row=i, column=j).value = copycell

# Loops through rows, and copies all from where blank rows are to be inserted
# and adds them to the new file, after the correct amount of blank rows
for i in range(0, mr + 1):
    for j in range(1, mc + 1):
        copycell = ws1.cell(row=i + rowNum, column=j).value
        ws2.cell(row=i + blankNum + rowNum, column=j).value = copycell

# Save new workbook
wb2.save('rowinserterProduceSales.xlsx')