"""Multiplication Table.
Create a multiplication table from a given number.
"""

import openpyxl
from openpyxl.styles import Font

numN = int(input('Please enter a number for the muliplication table: '))
# Could do it from command line, but I don't want to tbh...

# Opens an Excel workbook object, assigns the active sheet, and creates
# a bold font object
wb = openpyxl.Workbook()
sheet = wb.active
bold = Font(bold = True)

# Labels (relevant number) column headers with the relevant number
for htitle in range(1, numN + 1):
    sheet.cell(row=1, column=htitle + 1).value = htitle
    # Makes the headers bold!
    sheet.cell(row=1, column=htitle + 1).font = bold

# Labels (relevant number) rows with the relevant number
for rtitle in range(1, numN + 1):
    sheet.cell(row=rtitle + 1, column=1).value = rtitle
    # Makes the row label bold!
    sheet.cell(row=rtitle + 1, column=1).font = bold

# Nested loop to fill in each column for each row with the associated
# multiplied amount, except, obviously, the column and row labels
for rowNum in range(2, sheet.max_row + 1):
    rmult = sheet.cell(row=rowNum, column=1).value
    for colNum in range(2, sheet.max_column + 1):
        cmult = sheet.cell(row=1, column=colNum).value
        sheet.cell(row=rowNum, column=colNum).value = rmult * cmult

# Save the spreadsheet to the cwd.
wb.save('multiplicationTable.xlsx')